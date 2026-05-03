import json, io
import pandas as pd
import xmltodict, urllib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


DARK_BLUE = "1F3864"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREY = "F5F5F5"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED = "FCE4D6"
LIGHT_YELLOW = "FFF2CC"


def convert_xml_to_json(file_name: str):
    # if the file is in xml format
    with open(file_name, "r") as file:
        data = xmltodict.parse(file.read())
        data = data["root"]

        with open(file_name.replace(".xml", ".json"), "w") as file:
            json.dump(data, file)


def inspect(data: dict):
    # understanding the shape
    print(type(data))  # dict or list?
    print(len(data))  # how many records

    # if it's a list of dicts
    df = pd.json_normalize(data)
    print(df.shape)
    print(df.dtypes)
    print(df.head())


def write_to_excel(rows: list[dict], ws_title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = ws_title

    # writing the headers
    headers = list(rows[0].keys())
    ws.append(headers)

    # writing the rows
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    return wb


def apply_generic_style(wb: Workbook) -> Workbook:
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # adding the header style
    for col_num, value in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        make_header_cell(cell, value)

    ws.row_dimensions[1].height = 30

    # Alternating row colors
    white_fill = PatternFill(fill_type="solid", fgColor=WHITE)
    light_grey_fill = PatternFill(fill_type="solid", fgColor=LIGHT_GREY)

    for row in range(2, ws.max_row + 1):
        # checking for even and odd row number
        fill = white_fill if row % 2 == 0 else light_grey_fill

        # adding the row color change
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    # freezes the rows before A2 (the header row)
    # when scrolling down the header row always remain visible
    ws.freeze_panes = "A2"

    # using the header row as a reference for auto_filler function
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # auto size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Align center
            cell.alignment = center_alignment
    return wb


def make_header_cell(cell, text):
    """Style a single header cell: dark blue bg, white bold text, centered."""
    cell.value = text
    cell.font = Font(name="Noto Sans Lisu", bold=True, color=WHITE, size=10)
    cell.fill = PatternFill(fill_type="solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def img_parser(wb: Workbook, column_name: str):
    ws = wb.active

    # Extract headers and data from worksheet
    data_points = list(ws.values)
    if not data_points:
        return wb

    headers = data_points[0]
    data_rows = data_points[1:]

    # Map column name to column index and letter
    try:
        col_idx = list(headers).index(column_name) + 1
        col_letter = get_column_letter(col_idx)
    except ValueError:
        print(f"Column '{column_name}' not found")
        return wb

    # getting all the images
    for i, row in enumerate(data_rows, 1):
        url = row[col_idx - 1]

        # skipping if no url or not a string
        if pd.isna(url) or not isinstance(url, str) or not url.startswith("http"):
            continue

        try:
            # Download image into memory
            with urllib.request.urlopen(url) as response:
                img_data = response.read()

            img_file = io.BytesIO(img_data)
            img = Image(img_file)

            # Excel height is in points, width is roughly in characters
            ws.row_dimensions[i + 1].height = img.height * 0.75

            ws.add_image(img, f"{col_letter}{i + 1}")

            # Clear the URL text so only the image is visible
            ws.cell(row=i + 1, column=col_idx).value = ""

        except Exception as e:
            print(f"Error processing image {i}: {e}")

    return wb

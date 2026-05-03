from datetime import datetime
import json
import parser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def main():
    json_filename = "data/user_data.json"
    xlsx_filename = json_filename.replace(".json", ".xlsx")

    # loading the json file
    with open(json_filename, "r") as file:
        raw_json = json.load(file)

    rows = flatten_reservations(raw_json)

    if rows:
        wb = parser.write_to_excel(rows, ws_title="Users")
        wb = parser.img_parser(wb, "Picture")
        wb = parser.apply_generic_style(wb)

        # wb = apply_custom_style(wb)
        # wb = write_summary(wb, raw_json["payload"]["orders"], rows)
        wb.save(xlsx_filename)
    else:
        print("No rows to write")


def flatten_reservations(data: dict) -> list[dict]:
    try:
        rows = []
        for result in data["results"]:
            row = {
                "Picture": result["picture"]["thumbnail"],
                "Name": f"{result['name']['title']} {result['name']['first']} {result['name']['last']}",
                "Gender": result["gender"].title(),
                "Phone": result["phone"],
                "Cell": result["cell"],
                "Email": result["email"],
                "Location": f"{result['location']['street']['number']} {result['location']['street']['name']}",
                "City": result["location"]["city"],
                "State": result["location"]["state"],
                "Country": result["location"]["country"],
                "Postcode": result["location"]["postcode"],
                "Coordinates": f"{result['location']['coordinates']['latitude']}, {result['location']['coordinates']['longitude']}",
                "Timezone": result["location"]["timezone"]["offset"],
                "Username": result["login"]["username"],
                "Password": result["login"]["password"],
                "Date of Birth": datetime.fromisoformat(result["dob"]["date"]).strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
                "Age": result["dob"]["age"],
                "Registration Date": datetime.fromisoformat(
                    result["registered"]["date"]
                ).strftime("%Y-%m-%d %H:%M:%S"),
                "Account Age": result["registered"]["age"],
            }
            # deduplicationn
            if row not in rows:
                rows.append(row)
        return rows
    except Exception as error:
        print(error)
        return None


def apply_custom_style(wb: Workbook) -> Workbook:
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    STATUS_COLORS = {
        "Paid": parser.LIGHT_GREEN,
        "Pending": parser.LIGHT_YELLOW,
        "Refunded": parser.LIGHT_BLUE,
        "Failed": parser.LIGHT_RED,
    }
    color_column = headers.index("Payment Status") + 1

    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=color_column)
        status_value = status_cell.value

        # changing the color for each cell in status column
        if status_value in STATUS_COLORS:
            status_cell.fill = PatternFill(
                fill_type="solid", fgColor=STATUS_COLORS[status_value]
            )

    # Number formats
    currency_columns = ["Total Order", "Shipping Cost"]
    percent_column = []

    # adding the correct format
    for col_num, header in enumerate(headers, 1):
        if header in currency_columns:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_num).number_format = "$#,##0.00"
        elif header in percent_column:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_num).number_format = '0"%"'
    return wb


def write_summary(wb: Workbook, raw_json: dict, rows: list[dict]) -> Workbook:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    """Calculation Block - calculating all the data before writing"""
    total_orders = len(rows)
    total_revenue = sum(
        row["Order Total"] for row in rows if row["Payment Status"] != "Refunded"
    )

    items_sold = 0

    for order in raw_json:
        for item in order["items"]:
            items_sold += item["qty"]

    avg_order_value = round(total_revenue / total_orders, 2)

    order_stats = {}

    for row in rows:
        status = row["Order Status"]

        if status not in order_stats:
            order_stats[status] = 0
        order_stats[status] += 1

    item_stats = {}

    for order in raw_json:
        for item in order["items"]:
            item_name = item["name"]
            item_qty = item["qty"]

            if item_name not in item_stats:
                item_stats[item_name] = 0
            item_stats[item_name] += item_qty

    """Writing block - writing the summary stats"""
    for label, value, label_cell, value_cell in [
        ("Total Orders", total_orders, "B1", "B2"),
        ("Total Revenue", f"${total_revenue:,.2f}", "F1", "F2"),
        ("Items Sold", items_sold, "B4", "B5"),
        ("Avg Order Value", f"${avg_order_value:,.2f}", "F4", "F5"),
    ]:
        lc = ws[label_cell]
        vc = ws[value_cell]

        lc.value = label
        lc.font = Font(name="Noto Sans Lisu", bold=True, size=10, color=parser.WHITE)
        lc.fill = PatternFill("solid", fgColor=parser.DARK_BLUE)
        lc.alignment = Alignment(horizontal="center")
        lc.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        vc.value = value
        vc.font = Font(name="Bahnschrift", bold=True, size=18, color=parser.DARK_BLUE)
        vc.fill = PatternFill("solid", fgColor=parser.LIGHT_BLUE)
        vc.alignment = Alignment(horizontal="center")
        vc.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    status_headers = ["Order Status", "Orders"]
    # Starts at row 9 (rows 6-8 act as visual breathing room)
    status_start_row = 9

    # Write header row
    for column_index, header in enumerate(
        status_headers, 2
    ):  # start at column B (index 2)
        parser.make_header_cell(
            ws.cell(row=status_start_row, column=column_index), header
        )

    sorted_order = sorted(order_stats.items(), key=lambda x: x[1], reverse=True)

    for row_index, (status, stats) in enumerate(sorted_order, status_start_row + 1):
        bg = (
            parser.WHITE if row_index % 2 == 0 else parser.LIGHT_GREY
        )  # alternating rows

        data = [
            status,
            stats,
        ]
        for column_index, val in enumerate(data, 2):
            cell = ws.cell(row=row_index, column=column_index)
            cell.value = val
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Bahnschrift", size=10)
            cell.alignment = Alignment(
                horizontal="right" if column_index > 2 else "left"
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    item_headers = ["Item Name", "Quantity Sold"]
    item_start_row = 9

    for column_index, header in enumerate(item_headers, 2):
        parser.make_header_cell(
            ws.cell(row=item_start_row, column=3 + column_index), header
        )

    sorted_items = sorted(item_stats.items(), key=lambda x: x[1], reverse=True)

    for row_index, (name, stats) in enumerate(sorted_items, item_start_row + 1):
        bg = parser.WHITE if row_index % 2 == 0 else parser.LIGHT_GREY
        data = [name, stats]

        for column_index, val in enumerate(data, 2):
            cell = ws.cell(row=row_index, column=3 + column_index)
            cell.value = val
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Bahnschrift", size=10)
            cell.alignment = Alignment(
                horizontal="right" if column_index in (3, 4) else "left"
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    # auto size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    return wb


if __name__ == "__main__":
    main()
